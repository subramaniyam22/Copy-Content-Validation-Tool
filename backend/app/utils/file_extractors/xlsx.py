"""XLSX text extraction using openpyxl."""
from io import BytesIO

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


def extract_text_from_xlsx(file_bytes: bytes) -> str:
    """Extract text content from an XLSX file (all sheets, all cells)."""
    if load_workbook is None:
        return "[XLSX extraction requires openpyxl — install with: pip install openpyxl]"

    wb = load_workbook(BytesIO(file_bytes), read_only=True)
    text_parts = []
    for sheet in wb.worksheets:
        sheet_lines = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell) for cell in row if cell is not None]
            if cells:
                sheet_lines.append(" | ".join(cells))
        if sheet_lines:
            text_parts.append(f"--- Sheet: {sheet.title} ---\n" + "\n".join(sheet_lines))
    wb.close()
    return "\n\n".join(text_parts)
